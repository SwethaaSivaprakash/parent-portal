
import xlrd
import openpyxl
from flask import *
from openpyxl import load_workbook
from string import Template
import smtplib, ssl
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import glob#to list all files in directory
uname=""
coursename=""
total=0
CGPA=[]
app=Flask(__name__)
app.secret_key = "kannihilators"
ALLOWED_EXTENSIONS = {'xlsx'}
UPLOAD_FOLDER = 'D:\\auxano_connections\\uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
@app.errorhandler(404)
def page_not_found(name):
	with app.app_context():
		return render_template("404.html")
def get_contacts(filename,a1):
	names = []
	emails = []
	cname=[]
	try:
		print('step 1')
		wb = xlrd.open_workbook(filename) #opens he excel sheet
		sheet = wb.sheet_by_index(0) #gets the first sheet
		sheet.cell_value(1,1)#gets the first sheet containing names and mail
		total=sheet.nrows-1
		print(total)
		for i in range(sheet.nrows-1):
			names.append(sheet.cell_value(i+1,0))
			#print(names)	#get names
			emails.append(sheet.cell_value(i+1,1))
			#print(emails)	#get email
		print('name and email packed')
		#print(a1)
		if (a1=='CAT 1'):#if option selected is cat1 access cat1 sheet
			print(a1)
			sheet1=wb.sheet_by_name('CAT1')#open CAT 1sheet
			print(sheet1.ncols)
			for j in range(1,sheet1.ncols+1,2):
				#print(sheet1.cell_value(j,0))
				cname.append(sheet1.cell_value(0,j))
				print(cname)
			marks=[]
			attendance=[]
			for i in range(sheet1.nrows-1):
				marks.append(sheet1.cell_value(i+1,0))
			print(marks)
			for i in range(sheet1.nrows-1):
				marks.append(sheet1.cell_value(i+1,0))
			print(marks)
			#marksnattendance=[[0]*ncols]*(nrows-1)#create a two d marks and attendance list for every student ncols containing his marks
			#print(marksnattendance)
			for i in range(sheet1.nrows-1):	
				for j in range(sheet1.ncols):
					marksnattendance[j][i].append(sheet1.cell_value(i+1,j))#append marks n attendance
		elif (a1=='CAT 2'):
			sheet1=wb['CAT 2']
			for i in range(sheet1.nclos,2):
				cname.append(i)
				print(i)
			marksnattendance=[[0]*sheet1.ncols]*(sheet1.nrows-1)
			for i in range(sheet1.nrows-1):	
				for j in range(sheet1.ncols):
					marksnattendance[j][i].append(sheet1.cell_value(i+1,j))
		elif (a1=='CAT 3'):
			sheet1=wb['CAT 3']
			for i in range(sheet1.nclos,2):
				cname.append(i)
				print(i)
			marksnattendance=[[0]*sheet1.ncols]*(sheet1.nrows-1)
			for i in range(sheet1.nrows-1):	
				for j in range(sheet1.ncols):
					marksnattendance[j][i].append(sheet1.cell_value(i+1,j))
		else:#if terminal option is selected get open the terminal sheet and final grade sheet
			sheet1=wb['Terminal']
			for i in range(sheet1.nclos,2):
				cname.append(i)#to get course name which is in alternative cols
				print(i)
			marksnattendance=[[0]*sheet1.ncols]*(sheet1.nrows-1)
			for i in range(sheet1.nrows-1):	
				for j in range(sheet1.ncols):
					marksnattendance[j][i].append(sheet1.cell_value(i+1,j))
				CGPA=sheet1.cell_value(i+1,sheet1.ncols)#o get cgpa of each student which is the last column of the sheet
		if (names==[] or emails==[] or marks==[] or attendance==[] or cname==" "):
			raise Exception("Content not found")

		return names,emails,attendance,marks,cname,total
	except FileNotFoundError as e:
		with app.app_context():
			return render_template("FNF.html",errormsg=str(e))
	except Exception as e:
		with app.app_context():
			return render_template("Error.html",errormsg=str(e))
def read_template(filename):
	try:
		with open(filename, 'r', encoding='utf-8') as template_file:
			template_file_content = template_file.read()
			return Template(template_file_content)
	except FileNotFoundError as e:
		with app.app_context():
			return render_template("FNF.html",errormsg=str(e))
@app.route('/mail',methods=['GET', 'POST'])			
def sendmail(**nthng):
	try:
		if request.method == "POST":
			select = request.form.get('comp_select')
			a=str(select)
			print(a)
			f = request.files['file']  
			f1=f.filename
			print(f1)
			password = request.form["username"]
			with app.app_context():
				count=0
				names,emails,attendance,marks,cname,tot = get_contacts(f1,a)  # read contacts
				port = 465  # For SSL
				# Create a secure SSL context
				context = ssl.create_default_context()
				context = ssl.create_default_context()
				server = smtplib.SMTP('smtp.gmail.com')
				server.connect('smtp.gmail.com', '587')
				server.ehlo()
				server.starttls()
				server.login("auxanoconnections@gmail.com", password)
				marks=[]
				attendance=[]
				for i in range(len(marksnattendance)-1,2):
					attendance=marksnattendance[i+1]
					print(attendance)
				for i in range(len(marksnattendance)-1,2):
					marks=marksnattendance[i]
					print(marks)
				if (CGPA==[]):
					CGPA=['Not applicable']
				for name, email,attendance,marks in zip(names, emails,attendance,marks):
					msg = MIMEMultipart()       # create a message
					# add in the actual person name to the message template
					if (marks<25 and attendance<75.00):
						message.html=render_template(message.html,NAME=name.title(),CNAME=cname,ATTENDANCE=attendance,MARKS=marks,cgpa=CGPA,)
					elif (marks<25):
						message.html=render_template(message.html,NAME=name.title(),CNAME=cname,ATTENDANCE=attendance,MARKS=marks,cgpa=CGPA)
					elif (attendance<75.00):
						message.html=render_template(message.html,NAME=name.title(),CNAME=cname,ATTENDANCE=attendance,MARKS=marks,cgpa=CGPA)
					else:
						message.html=render_template(message.html,NAME=name.title(),CNAME=cname,ATTENDANCE=attendance,MARKS=marks,cgpa=CGPA)
					# setup the parameters of the message
					msg['From']="auxanoconnections@gmail.com"
					msg['To']=email
					msg['Subject']="This is TEST"
					# add in the message body
					msg.attach(MIMEText(message, 'plain'))
					# send the message via the server set up earlier.
					server.send_message(msg)
					count+=1
					del msg	 
				server.quit()
				print(count,tot)
				if (count==tot):
					return render_template("Success.html")
				else:
					return "Mail not sent"
		return render_template("mail1.html")	
	except Exception as  e:
		with app.app_context():
			return render_template("Error.html",errormsg=str(e))			
@app.route('/upload/<filename>')
def uploaded_file(filename):
    return send_from_directory(os.path.dirname(UPLOAD_FOLDER),filename)
#To check the file format
def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS
#Main page for uploading
@app.route("/uploadfile", methods=['GET', 'POST'])
def uploadfile():
	print("Here")
	if request.method == 'POST':
        #print(request.files['file'])
		file = request.files['file']
		if file.filename == '':
			flash('No selected file')
			return redirect(request.url)
		if file and allowed_file(file.filename):
			file.save(os.path.join(app.config['UPLOAD_FOLDER'],file.filename))
			append_data=filename.split('_')#to split file name 
			coursename=append_data[0]#to get Coursename from filename
			file_to_append=append_data[1]#to get exam type from filename
			file_name=filename[2:5]	#to find the year of file
			files_in_dir=glob.glob("/D:/auxano_connections/*.xlsx")
			for filen in files_in_dir:
				if file_name in filen:
					df=pd.open_workbook(filen)
					df1=pd.open_workbook(filename)
					sheet1=df1[file_to_append]#open the sheet in input file
					sheet=df[file_to_append]#open sheet in std file
					total=sheet.nrows-1
					cols = "A"
					data=[]
					count=1
					row=count
					c=str(row)
					cell=cols+c
					print(cell)
					for i in range(sheet.nrows-1):
						cell_val1=sheet1.cell_value(i+1,2)
						cell_val2=sheet1.cell_value(i+1,5)
						while(sheet1.cell(row=count,column=1)!=" "):
							y = str(row)
							cell = cols + y
							count+=1
							row=count
						for j in range(1,ncols,2):
							while(sheet.cell(row=count,column=j!=" ")):
								c1=sheet.cell(row=count,column=j)
								c1.value(cell_val1)
								c2=sheet.cell(row=count,column=j+1)
								c2.value(cell_val2)
			return redirect(url_for('uploaded_file',filename=file.filename))
            #data_xls = pd.read_excel(file)
            #return data_xls.to_html()
		else:
			submit_name = file.filename
            #if '.' in submit_name and submit_name.rsplit('.', 1)[1] in ALLOWED_EXTENSIONS:
            #    filename = secure_filename(submit_name)
            #    form.file_upload.data.save('uploads/' + filename)
             #   return redirect('home')
        #else:
			flash('File (%s) is not an accepted format' % submit_name)
			print (submit_name)
	return render_template('upload.html')
@app.route("/export", methods=['GET'])
def export_records():
    return 
@app.route('/suggestions', methods=['GET', 'POST'])
def suggestions(**nthng):
	if request.method == "POST":
		mail_id = request.form["text"]
		password = request.form["text1"]
		message = request.form["text2"]
		port = 465  # For SSL
		# Create a secure SSL context
		context = ssl.create_default_context()
		context = ssl.create_default_context()
		msg = MIMEMultipart()       # create a message
		msg['From']=mail_id
		msg['To']="auxanoconnections@gmail.com"
		msg['Subject']="Suggestions->Auxano_connections"
		server = smtplib.SMTP('smtp.gmail.com')
		server.connect('smtp.gmail.com', '587')
		server.starttls()
		server.login(msg['From'], password)
		# add in the message body
		msg.attach(MIMEText(message, 'plain'))
		# send the message via the server set up earlier.
		server.sendmail(msg['From'],msg['To'],msg.as_string())
		server.quit()
	return render_template('suggestions.html')
@app.route('/',methods=['GET', 'POST'])
def login():
	try:
		count=0
		if request.method == 'POST':
			uname = request.form['nm']
			password=request.form['pw']
			print(password)
			with app.app_context(): 
				df = (r"D:\auxano_connections\Faculty.xlsx")
				wb = xlrd.open_workbook(df) #opens the excel sheet
				sheet = wb.sheet_by_index(0) #gets the first sheet
				for x in range(sheet.nrows-1):
					un=sheet.cell_value(x+1,5)
					passw=sheet.cell_value(x+1,6)
					print(passw)
					if (uname == un and password==passw):
						count+=1
						cid=sheet.cell_value(x+1,4).split(',')
						return render_template('Course.html',cid=cid)
				if count==0:
					return render_template('UsernotFound.html')
	except FileNotFoundError:
		print("File not found!")
	except Exception as e:
		print(e)
	if "course" in request.form:
		return render_template('option.html')
	elif "mail" in request.form:	
		return render_template('mail1.html')
	
	elif "upload" in request.form:	
		return render_template('upload.html')		
	else:
		return render_template('login.html')

if __name__=='__main__':
	app.debug=True
	app.run()